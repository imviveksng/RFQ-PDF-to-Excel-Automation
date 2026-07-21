from openpyxl import load_workbook
from pathlib import Path
from config import EXCEL_FILE, SHEET_NAME

from mappings import COLUMN_MAPPING


def write_to_excel(data):
    print("Excel Path:", EXCEL_FILE)
    print("Exists:", Path(EXCEL_FILE).exists())

    workbook = load_workbook(EXCEL_FILE)

    sheet = workbook[SHEET_NAME]

    next_row = sheet.max_row + 1

    for field, column in COLUMN_MAPPING.items():

        sheet[f"{column}{next_row}"] = data.get(field, "")

    workbook.save(str(EXCEL_FILE))

    workbook.close()

    print(f"✓ Saved to Row {next_row}")